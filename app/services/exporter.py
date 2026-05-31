"""Export test cases to Excel (XLSX) and CSV formats."""

import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXPORT_FIELDS = [
    "case_id", "module", "sub_module", "title", "preconditions",
    "steps", "expected_result", "keywords", "priority", "category",
    "applicable_phase", "description", "reviewer", "test_method",
    "estimated_time", "notes", "test_frequency", "test_level", "duration",
    "tags", "review_status", "review_comment", "execution_status",
]

EXPORT_HEADERS = [
    "用例编号", "三级模块名称", "子模块", "用例标题", "前置条件",
    "测试步骤", "预期结果", "关键字/标签", "优先级", "用例类型",
    "适用阶段", "用例说明", "由谁评审", "测试方法",
    "预计执行时间", "其他/备注", "测试频率", "测试级别", "时长",
    "标签", "评审状态", "评审意见", "执行状态",
]

def _normalize_cases(test_cases: list[dict]) -> list[dict]:
    result = []
    for tc in test_cases:
        normalized = {}
        for f in EXPORT_FIELDS:
            normalized[f] = tc.get(f, "")
        result.append(normalized)
    return result


def export_to_xlsx(test_cases: list[dict]) -> io.BytesIO:
    cases = _normalize_cases(test_cases)
    if not cases:
        cases = [dict.fromkeys(EXPORT_FIELDS, "")]

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Arial", size=9)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    p0_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    p1_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

    # Write headers
    for col_idx, header in enumerate(EXPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    # Write data rows
    for row_idx, tc in enumerate(cases, 2):
        for col_idx, field in enumerate(EXPORT_FIELDS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=tc.get(field, ""))
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border
            # Color-code priority columns
            if field == "priority":
                if tc.get(field) == "P0":
                    cell.fill = p0_fill
                    cell.font = Font(name="Arial", size=9, bold=True, color="CC0000")
                elif tc.get(field) == "P1":
                    cell.fill = p1_fill

    # Column widths
    col_widths = [12, 16, 14, 35, 28, 38, 32, 16, 8, 10, 12, 28, 10, 12, 10, 16, 10, 12, 10, 16, 12, 28, 12]
    for col_idx, width in enumerate(col_widths, 1):
        if col_idx <= len(EXPORT_FIELDS):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze top row and auto-filter
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(EXPORT_FIELDS))
    ws.auto_filter.ref = f"A1:{last_col}{len(cases) + 1}"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_to_csv(test_cases: list[dict]) -> io.BytesIO:
    import codecs
    cases = _normalize_cases(test_cases)

    buffer = io.BytesIO()
    buffer.write(codecs.BOM_UTF8)
    wrapper = io.TextIOWrapper(buffer, "utf-8", newline="")

    # Write headers row using Chinese headers
    header_dict = dict(zip(EXPORT_FIELDS, EXPORT_HEADERS))
    writer = csv.DictWriter(wrapper, fieldnames=EXPORT_FIELDS, extrasaction="ignore")
    writer.writerow(header_dict)
    for tc in cases:
        writer.writerow(tc)

    wrapper.detach()
    buffer.seek(0)
    return buffer
